"""Advanced analytics module for trend analysis, anomaly detection, and correlations.

This module provides comprehensive analytics capabilities for test results,
including trend analysis, anomaly detection, correlation analysis, and
data export functionality.
"""

from __future__ import annotations

import csv
import importlib.util
import io
import sysconfig
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

if TYPE_CHECKING:
    from types import ModuleType


def _get_stdlib_statistics() -> ModuleType:
    """Get the standard library statistics module.

    There's a local atp.statistics module that can shadow the stdlib statistics.
    This function ensures we get the stdlib version.
    """
    spec = importlib.util.find_spec("statistics")
    if spec is None or spec.origin is None:
        raise ImportError("Cannot find stdlib statistics module")

    # Check if this is the stdlib version (not a local shadowing module)
    if "atp" in spec.origin or "tests" in spec.origin:
        # Need to load from stdlib path
        stdlib_path = sysconfig.get_paths()["stdlib"]
        spec = importlib.util.spec_from_file_location(
            "statistics", f"{stdlib_path}/statistics.py"
        )
        if spec is None or spec.loader is None:
            raise ImportError("Cannot find stdlib statistics module")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    else:
        import statistics

        return statistics


_stdlib_stats = _get_stdlib_statistics()
stats_mean = _stdlib_stats.mean
stats_stdev = _stdlib_stats.stdev


class TrendDirection(str, Enum):
    """Direction of a trend."""

    IMPROVING = "improving"
    DECLINING = "declining"
    STABLE = "stable"
    INSUFFICIENT_DATA = "insufficient_data"


class AnomalyType(str, Enum):
    """Type of detected anomaly."""

    SCORE_SPIKE = "score_spike"
    SCORE_DROP = "score_drop"
    DURATION_SPIKE = "duration_spike"
    ERROR_RATE_SPIKE = "error_rate_spike"
    COST_SPIKE = "cost_spike"


class CorrelationStrength(str, Enum):
    """Strength of a correlation."""

    STRONG_POSITIVE = "strong_positive"
    MODERATE_POSITIVE = "moderate_positive"
    WEAK_POSITIVE = "weak_positive"
    NO_CORRELATION = "no_correlation"
    WEAK_NEGATIVE = "weak_negative"
    MODERATE_NEGATIVE = "moderate_negative"
    STRONG_NEGATIVE = "strong_negative"


# ==================== Pydantic Models ====================


class TrendDataPoint(BaseModel):
    """Single data point in a trend analysis."""

    model_config = ConfigDict(from_attributes=True)

    date: str
    value: float
    count: int = 1


class TrendAnalysisResult(BaseModel):
    """Result of trend analysis."""

    model_config = ConfigDict(from_attributes=True)

    metric: str
    direction: TrendDirection
    change_percent: float
    start_value: float
    end_value: float
    average_value: float
    std_deviation: float
    data_points: list[TrendDataPoint]
    period_days: int
    confidence: float = Field(ge=0.0, le=1.0)


class AnomalyResult(BaseModel):
    """Detected anomaly in test results."""

    model_config = ConfigDict(from_attributes=True)

    anomaly_type: AnomalyType
    timestamp: datetime
    metric_name: str
    expected_value: float
    actual_value: float
    deviation_sigma: float
    test_id: str | None = None
    suite_id: str | None = None
    agent_name: str | None = None
    severity: str = "medium"  # low, medium, high
    details: dict[str, Any] = Field(default_factory=dict)


class CorrelationResult(BaseModel):
    """Correlation between two factors."""

    model_config = ConfigDict(from_attributes=True)

    factor_x: str
    factor_y: str
    correlation_coefficient: float = Field(ge=-1.0, le=1.0)
    strength: CorrelationStrength
    sample_size: int
    p_value: float | None = None
    details: dict[str, Any] = Field(default_factory=dict)


class ScoreTrendResponse(BaseModel):
    """Response for score trends API."""

    model_config = ConfigDict(from_attributes=True)

    suite_name: str | None = None
    agent_name: str | None = None
    trends: list[TrendAnalysisResult]
    period_start: datetime
    period_end: datetime


class AnomalyDetectionResponse(BaseModel):
    """Response for anomaly detection API."""

    model_config = ConfigDict(from_attributes=True)

    anomalies: list[AnomalyResult]
    total_records_analyzed: int
    anomaly_rate: float
    period_start: datetime
    period_end: datetime


class CorrelationAnalysisResponse(BaseModel):
    """Response for correlation analysis API."""

    model_config = ConfigDict(from_attributes=True)

    correlations: list[CorrelationResult]
    sample_size: int
    factors_analyzed: list[str]


class ExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    EXCEL = "excel"


class ScheduledReportFrequency(str, Enum):
    """Frequency for scheduled reports."""

    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"


class ScheduledReportConfig(BaseModel):
    """Configuration for a scheduled report."""

    model_config = ConfigDict(from_attributes=True)

    id: int | None = None
    name: str = Field(min_length=1, max_length=100)
    frequency: ScheduledReportFrequency
    recipients: list[str] = Field(default_factory=list)
    include_trends: bool = True
    include_anomalies: bool = True
    include_correlations: bool = False
    filters: dict[str, Any] = Field(default_factory=dict)
    is_active: bool = True
    last_run: datetime | None = None
    next_run: datetime | None = None
    created_at: datetime | None = None


# ==================== Internal Data Classes ====================


@dataclass
class MetricData:
    """Internal data class for metric calculations."""

    values: list[float] = field(default_factory=list)
    timestamps: list[datetime] = field(default_factory=list)
    ids: list[int] = field(default_factory=list)


# ==================== Analytics Service ====================


class AdvancedAnalyticsService:
    """Service for advanced analytics operations.

    Provides trend analysis, anomaly detection, correlation analysis,
    and data export functionality for test results.
    """

    # Anomaly detection thresholds (in standard deviations)
    ANOMALY_THRESHOLD_HIGH = 3.0
    ANOMALY_THRESHOLD_MEDIUM = 2.0
    ANOMALY_THRESHOLD_LOW = 1.5

    # Minimum data points for reliable analysis
    MIN_DATA_POINTS_TREND = 3
    MIN_DATA_POINTS_ANOMALY = 10
    MIN_DATA_POINTS_CORRELATION = 5

    def __init__(self, session: AsyncSession):
        """Initialize the analytics service.

        Args:
            session: SQLAlchemy async session.
        """
        self._session = session

    async def analyze_score_trends(
        self,
        *,
        suite_name: str | None = None,
        agent_name: str | None = None,
        test_id: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        metrics: list[str] | None = None,
    ) -> ScoreTrendResponse:
        """Analyze score trends over time.

        Args:
            suite_name: Optional filter by suite name.
            agent_name: Optional filter by agent name.
            test_id: Optional filter by test ID.
            start_date: Start of analysis period.
            end_date: End of analysis period.
            metrics: Metrics to analyze (default: score, success_rate, duration).

        Returns:
            ScoreTrendResponse with trend analysis results.
        """
        from atp.dashboard.models import Agent, SuiteExecution, TestExecution

        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=30)
        if metrics is None:
            metrics = ["score", "success_rate", "duration"]

        # Build query for test executions
        stmt = (
            select(TestExecution)
            .join(SuiteExecution)
            .options(selectinload(TestExecution.suite_execution))
            .where(TestExecution.started_at >= start_date)
            .where(TestExecution.started_at <= end_date)
            .order_by(TestExecution.started_at)
        )

        if suite_name:
            stmt = stmt.where(SuiteExecution.suite_name == suite_name)
        if agent_name:
            stmt = stmt.join(Agent).where(Agent.name == agent_name)
        if test_id:
            stmt = stmt.where(TestExecution.test_id == test_id)

        result = await self._session.execute(stmt)
        executions = list(result.scalars().all())

        trends: list[TrendAnalysisResult] = []

        for metric in metrics:
            trend = self._calculate_trend(executions, metric)
            if trend:
                trends.append(trend)

        return ScoreTrendResponse(
            suite_name=suite_name,
            agent_name=agent_name,
            trends=trends,
            period_start=start_date,
            period_end=end_date,
        )

    def _calculate_trend(
        self, executions: list, metric: str
    ) -> TrendAnalysisResult | None:
        """Calculate trend for a specific metric.

        Args:
            executions: List of test executions.
            metric: Metric to analyze.

        Returns:
            TrendAnalysisResult or None if insufficient data.
        """
        # Extract values based on metric
        data_by_date: dict[str, list[float]] = {}

        for exec in executions:
            date_str = exec.started_at.strftime("%Y-%m-%d")

            if metric == "score":
                value = exec.score
            elif metric == "success_rate":
                value = 1.0 if exec.success else 0.0
            elif metric == "duration":
                value = exec.duration_seconds
            else:
                continue

            if value is not None:
                if date_str not in data_by_date:
                    data_by_date[date_str] = []
                data_by_date[date_str].append(value)

        if len(data_by_date) < self.MIN_DATA_POINTS_TREND:
            return TrendAnalysisResult(
                metric=metric,
                direction=TrendDirection.INSUFFICIENT_DATA,
                change_percent=0.0,
                start_value=0.0,
                end_value=0.0,
                average_value=0.0,
                std_deviation=0.0,
                data_points=[],
                period_days=0,
                confidence=0.0,
            )

        # Calculate daily averages
        data_points: list[TrendDataPoint] = []
        all_values: list[float] = []

        for date_str in sorted(data_by_date.keys()):
            values = data_by_date[date_str]
            avg = stats_mean(values)
            data_points.append(
                TrendDataPoint(date=date_str, value=avg, count=len(values))
            )
            all_values.extend(values)

        if not data_points:
            return None

        # Calculate statistics
        start_value = data_points[0].value
        end_value = data_points[-1].value
        average_value = stats_mean(all_values)
        std_dev = stats_stdev(all_values) if len(all_values) > 1 else 0.0

        # Calculate percent change
        if start_value != 0:
            change_percent = ((end_value - start_value) / abs(start_value)) * 100
        else:
            change_percent = 0.0 if end_value == 0 else 100.0

        # Determine direction
        threshold = 5.0  # 5% threshold for determining direction
        if change_percent > threshold:
            direction = TrendDirection.IMPROVING
        elif change_percent < -threshold:
            direction = TrendDirection.DECLINING
        else:
            direction = TrendDirection.STABLE

        # Calculate confidence based on sample size and consistency
        sample_size = len(all_values)
        confidence = min(1.0, sample_size / 50)  # Full confidence at 50+ samples

        # Adjust for duration metric (lower is better)
        if metric == "duration" and direction != TrendDirection.STABLE:
            direction = (
                TrendDirection.IMPROVING
                if change_percent < -threshold
                else TrendDirection.DECLINING
            )

        return TrendAnalysisResult(
            metric=metric,
            direction=direction,
            change_percent=round(change_percent, 2),
            start_value=round(start_value, 4),
            end_value=round(end_value, 4),
            average_value=round(average_value, 4),
            std_deviation=round(std_dev, 4),
            data_points=data_points,
            period_days=len(data_points),
            confidence=round(confidence, 2),
        )

    async def detect_anomalies(
        self,
        *,
        suite_name: str | None = None,
        agent_name: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        metrics: list[str] | None = None,
        sensitivity: str = "medium",
    ) -> AnomalyDetectionResponse:
        """Detect anomalies in test results.

        Uses statistical analysis (z-score) to identify unusual results
        that deviate significantly from the baseline.

        Args:
            suite_name: Optional filter by suite name.
            agent_name: Optional filter by agent name.
            start_date: Start of analysis period.
            end_date: End of analysis period.
            metrics: Metrics to analyze for anomalies.
            sensitivity: Detection sensitivity (low, medium, high).

        Returns:
            AnomalyDetectionResponse with detected anomalies.
        """
        from atp.dashboard.models import Agent, SuiteExecution, TestExecution

        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=30)
        if metrics is None:
            metrics = ["score", "duration"]

        # Set threshold based on sensitivity
        threshold_map = {
            "low": self.ANOMALY_THRESHOLD_HIGH,
            "medium": self.ANOMALY_THRESHOLD_MEDIUM,
            "high": self.ANOMALY_THRESHOLD_LOW,
        }
        threshold = threshold_map.get(sensitivity, self.ANOMALY_THRESHOLD_MEDIUM)

        # Query executions
        stmt = (
            select(TestExecution)
            .join(SuiteExecution)
            .options(selectinload(TestExecution.suite_execution))
            .where(TestExecution.started_at >= start_date)
            .where(TestExecution.started_at <= end_date)
            .order_by(TestExecution.started_at)
        )

        if suite_name:
            stmt = stmt.where(SuiteExecution.suite_name == suite_name)
        if agent_name:
            stmt = stmt.join(Agent).where(Agent.name == agent_name)

        result = await self._session.execute(stmt)
        executions = list(result.scalars().all())

        anomalies: list[AnomalyResult] = []

        for metric in metrics:
            metric_anomalies = self._detect_metric_anomalies(
                executions, metric, threshold
            )
            anomalies.extend(metric_anomalies)

        # Sort by severity and timestamp
        anomalies.sort(
            key=lambda a: (
                {"high": 0, "medium": 1, "low": 2}.get(a.severity, 1),
                a.timestamp,
            )
        )

        anomaly_rate = len(anomalies) / len(executions) if executions else 0.0

        return AnomalyDetectionResponse(
            anomalies=anomalies,
            total_records_analyzed=len(executions),
            anomaly_rate=round(anomaly_rate, 4),
            period_start=start_date,
            period_end=end_date,
        )

    def _detect_metric_anomalies(
        self, executions: list, metric: str, threshold: float
    ) -> list[AnomalyResult]:
        """Detect anomalies for a specific metric.

        Args:
            executions: List of test executions.
            metric: Metric to analyze.
            threshold: Z-score threshold for anomaly detection.

        Returns:
            List of detected anomalies.
        """
        # Extract values
        values: list[tuple[datetime, float, Any]] = []

        for exec in executions:
            if metric == "score":
                value = exec.score
            elif metric == "duration":
                value = exec.duration_seconds
            else:
                continue

            if value is not None:
                values.append((exec.started_at, value, exec))

        if len(values) < self.MIN_DATA_POINTS_ANOMALY:
            return []

        # Calculate statistics
        numeric_values = [v[1] for v in values]
        mean = stats_mean(numeric_values)
        std_dev = stats_stdev(numeric_values)

        if std_dev == 0:
            return []

        anomalies: list[AnomalyResult] = []

        for timestamp, value, exec in values:
            z_score = (value - mean) / std_dev

            if abs(z_score) >= threshold:
                # Determine severity
                if abs(z_score) >= self.ANOMALY_THRESHOLD_HIGH:
                    severity = "high"
                elif abs(z_score) >= self.ANOMALY_THRESHOLD_MEDIUM:
                    severity = "medium"
                else:
                    severity = "low"

                # Determine anomaly type
                if metric == "score":
                    if z_score < 0:
                        anomaly_type = AnomalyType.SCORE_DROP
                    else:
                        anomaly_type = AnomalyType.SCORE_SPIKE
                else:
                    anomaly_type = AnomalyType.DURATION_SPIKE

                anomalies.append(
                    AnomalyResult(
                        anomaly_type=anomaly_type,
                        timestamp=timestamp,
                        metric_name=metric,
                        expected_value=round(float(mean), 4),
                        actual_value=round(float(value), 4),
                        deviation_sigma=round(abs(float(z_score)), 2),
                        test_id=exec.test_id,
                        suite_id=str(exec.suite_execution_id),
                        severity=severity,
                        details={
                            "mean": round(float(mean), 4),
                            "std_dev": round(float(std_dev), 4),
                            "z_score": round(float(z_score), 2),
                        },
                    )
                )

        return anomalies

    async def analyze_correlations(
        self,
        *,
        suite_name: str | None = None,
        agent_name: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        factors: list[str] | None = None,
    ) -> CorrelationAnalysisResponse:
        """Analyze correlations between different factors and scores.

        Calculates Pearson correlation coefficients between factors like
        duration, token usage, tool calls, and test scores.

        Args:
            suite_name: Optional filter by suite name.
            agent_name: Optional filter by agent name.
            start_date: Start of analysis period.
            end_date: End of analysis period.
            factors: Factors to analyze (default: all available).

        Returns:
            CorrelationAnalysisResponse with correlation results.
        """
        from atp.dashboard.models import (
            Agent,
            SuiteExecution,
            TestExecution,
        )

        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=30)
        if factors is None:
            factors = ["duration", "total_tokens", "tool_calls", "llm_calls"]

        # Query test executions with run results
        stmt = (
            select(TestExecution)
            .join(SuiteExecution)
            .options(selectinload(TestExecution.run_results))
            .where(TestExecution.started_at >= start_date)
            .where(TestExecution.started_at <= end_date)
            .where(TestExecution.score.isnot(None))
        )

        if suite_name:
            stmt = stmt.where(SuiteExecution.suite_name == suite_name)
        if agent_name:
            stmt = stmt.join(Agent).where(Agent.name == agent_name)

        result = await self._session.execute(stmt)
        executions = list(result.scalars().all())

        if len(executions) < self.MIN_DATA_POINTS_CORRELATION:
            return CorrelationAnalysisResponse(
                correlations=[],
                sample_size=len(executions),
                factors_analyzed=factors,
            )

        # Extract data for correlation analysis
        scores: list[float] = []
        factor_data: dict[str, list[float]] = {f: [] for f in factors}

        for exec in executions:
            if exec.score is None:
                continue

            scores.append(exec.score)

            # Aggregate run result data
            total_tokens = 0
            tool_calls = 0
            llm_calls = 0

            for run in exec.run_results:
                total_tokens += (run.input_tokens or 0) + (run.output_tokens or 0)
                tool_calls += run.tool_calls or 0
                llm_calls += run.llm_calls or 0

            if "duration" in factor_data:
                factor_data["duration"].append(exec.duration_seconds or 0)
            if "total_tokens" in factor_data:
                factor_data["total_tokens"].append(total_tokens)
            if "tool_calls" in factor_data:
                factor_data["tool_calls"].append(tool_calls)
            if "llm_calls" in factor_data:
                factor_data["llm_calls"].append(llm_calls)

        correlations: list[CorrelationResult] = []

        for factor in factors:
            if factor not in factor_data:
                continue

            factor_values = factor_data[factor]
            if len(factor_values) != len(scores):
                continue

            correlation = self._calculate_correlation(factor_values, scores)
            if correlation is not None:
                correlations.append(
                    CorrelationResult(
                        factor_x=factor,
                        factor_y="score",
                        correlation_coefficient=round(correlation, 4),
                        strength=self._classify_correlation(correlation),
                        sample_size=len(scores),
                        details={"factor_mean": round(stats_mean(factor_values), 2)},
                    )
                )

        # Sort by absolute correlation strength
        correlations.sort(key=lambda c: abs(c.correlation_coefficient), reverse=True)

        return CorrelationAnalysisResponse(
            correlations=correlations,
            sample_size=len(scores),
            factors_analyzed=factors,
        )

    def _calculate_correlation(self, x: list[float], y: list[float]) -> float | None:
        """Calculate Pearson correlation coefficient.

        Args:
            x: First variable values.
            y: Second variable values.

        Returns:
            Correlation coefficient (-1 to 1) or None if cannot calculate.
        """
        if len(x) != len(y) or len(x) < 2:
            return None

        n = len(x)
        mean_x = stats_mean(x)
        mean_y = stats_mean(y)

        # Calculate Pearson correlation using the formula:
        # r = Σ((xi - mean_x)(yi - mean_y)) / sqrt(Σ(xi - mean_x)² * Σ(yi - mean_y)²)
        sum_xy = sum((x[i] - mean_x) * (y[i] - mean_y) for i in range(n))
        sum_xx = sum((x[i] - mean_x) ** 2 for i in range(n))
        sum_yy = sum((y[i] - mean_y) ** 2 for i in range(n))

        if sum_xx == 0 or sum_yy == 0:
            return None

        return sum_xy / (sum_xx * sum_yy) ** 0.5

    def _classify_correlation(self, r: float) -> CorrelationStrength:
        """Classify correlation strength.

        Args:
            r: Correlation coefficient.

        Returns:
            CorrelationStrength classification.
        """
        abs_r = abs(r)

        if abs_r >= 0.7:
            return (
                CorrelationStrength.STRONG_POSITIVE
                if r > 0
                else CorrelationStrength.STRONG_NEGATIVE
            )
        elif abs_r >= 0.4:
            return (
                CorrelationStrength.MODERATE_POSITIVE
                if r > 0
                else CorrelationStrength.MODERATE_NEGATIVE
            )
        elif abs_r >= 0.2:
            return (
                CorrelationStrength.WEAK_POSITIVE
                if r > 0
                else CorrelationStrength.WEAK_NEGATIVE
            )
        else:
            return CorrelationStrength.NO_CORRELATION

    async def export_to_csv(
        self,
        *,
        suite_name: str | None = None,
        agent_name: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        include_runs: bool = False,
    ) -> str:
        """Export test results to CSV format.

        Args:
            suite_name: Optional filter by suite name.
            agent_name: Optional filter by agent name.
            start_date: Start of export period.
            end_date: End of export period.
            include_runs: Whether to include individual run details.

        Returns:
            CSV content as string.
        """
        from atp.dashboard.models import (
            Agent,
            SuiteExecution,
            TestExecution,
        )

        if end_date is None:
            end_date = datetime.now()
        if start_date is None:
            start_date = end_date - timedelta(days=30)

        # Query test executions
        stmt = (
            select(TestExecution)
            .join(SuiteExecution)
            .options(
                selectinload(TestExecution.suite_execution).selectinload(
                    SuiteExecution.agent
                ),
                selectinload(TestExecution.run_results),
            )
            .where(TestExecution.started_at >= start_date)
            .where(TestExecution.started_at <= end_date)
            .order_by(TestExecution.started_at)
        )

        if suite_name:
            stmt = stmt.where(SuiteExecution.suite_name == suite_name)
        if agent_name:
            stmt = stmt.join(Agent).where(Agent.name == agent_name)

        result = await self._session.execute(stmt)
        executions = list(result.scalars().all())

        # Build CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        headers = [
            "timestamp",
            "suite_name",
            "agent_name",
            "test_id",
            "test_name",
            "success",
            "score",
            "duration_seconds",
            "total_runs",
            "successful_runs",
            "status",
        ]

        if include_runs:
            headers.extend(
                [
                    "run_number",
                    "run_success",
                    "run_duration",
                    "input_tokens",
                    "output_tokens",
                    "tool_calls",
                    "llm_calls",
                    "cost_usd",
                ]
            )

        writer.writerow(headers)

        # Write data
        for exec in executions:
            suite = exec.suite_execution
            agent = suite.agent if suite else None

            base_row = [
                exec.started_at.isoformat() if exec.started_at else "",
                suite.suite_name if suite else "",
                agent.name if agent else "",
                exec.test_id,
                exec.test_name,
                exec.success,
                exec.score,
                exec.duration_seconds,
                exec.total_runs,
                exec.successful_runs,
                exec.status,
            ]

            if include_runs and exec.run_results:
                for run in exec.run_results:
                    row = base_row + [
                        run.run_number,
                        run.success,
                        run.duration_seconds,
                        run.input_tokens,
                        run.output_tokens,
                        run.tool_calls,
                        run.llm_calls,
                        run.cost_usd,
                    ]
                    writer.writerow(row)
            else:
                writer.writerow(base_row)

        return output.getvalue()

    async def export_to_excel(
        self,
        *,
        suite_name: str | None = None,
        agent_name: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        include_trends: bool = True,
        include_anomalies: bool = True,
    ) -> bytes:
        """Export comprehensive analytics to Excel format.

        Creates a multi-sheet Excel workbook with:
        - Test Results sheet
        - Trends sheet (optional)
        - Anomalies sheet (optional)

        Args:
            suite_name: Optional filter by suite name.
            agent_name: Optional filter by agent name.
            start_date: Start of export period.
            end_date: End of export period.
            include_trends: Include trends analysis sheet.
            include_anomalies: Include anomalies sheet.

        Returns:
            Excel file content as bytes.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: uv add openpyxl"
            )

        workbook = openpyxl.Workbook()

        # Get CSV data for results sheet
        csv_data = await self.export_to_csv(
            suite_name=suite_name,
            agent_name=agent_name,
            start_date=start_date,
            end_date=end_date,
            include_runs=True,
        )

        # Create Results sheet - workbook.active is always a Worksheet
        results_sheet = workbook.active
        if results_sheet is None:
            results_sheet = workbook.create_sheet("Test Results")
        else:
            results_sheet.title = "Test Results"

        csv_reader = csv.reader(io.StringIO(csv_data))
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for row_idx, row in enumerate(csv_reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = results_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.fill = header_fill  # type: ignore[assignment]
                    cell.font = header_font  # type: ignore[assignment]

        # Add Trends sheet if requested
        if include_trends:
            trends_response = await self.analyze_score_trends(
                suite_name=suite_name,
                agent_name=agent_name,
                start_date=start_date,
                end_date=end_date,
            )

            trends_sheet = workbook.create_sheet("Trends")
            trends_headers = [
                "Metric",
                "Direction",
                "Change %",
                "Start Value",
                "End Value",
                "Average",
                "Std Dev",
                "Confidence",
            ]

            for col_idx, header in enumerate(trends_headers, 1):
                cell = trends_sheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, trend in enumerate(trends_response.trends, 2):
                trends_sheet.cell(row=row_idx, column=1, value=trend.metric)
                trends_sheet.cell(row=row_idx, column=2, value=trend.direction.value)
                trends_sheet.cell(row=row_idx, column=3, value=trend.change_percent)
                trends_sheet.cell(row=row_idx, column=4, value=trend.start_value)
                trends_sheet.cell(row=row_idx, column=5, value=trend.end_value)
                trends_sheet.cell(row=row_idx, column=6, value=trend.average_value)
                trends_sheet.cell(row=row_idx, column=7, value=trend.std_deviation)
                trends_sheet.cell(row=row_idx, column=8, value=trend.confidence)

        # Add Anomalies sheet if requested
        if include_anomalies:
            anomalies_response = await self.detect_anomalies(
                suite_name=suite_name,
                agent_name=agent_name,
                start_date=start_date,
                end_date=end_date,
            )

            anomalies_sheet = workbook.create_sheet("Anomalies")
            anomaly_headers = [
                "Timestamp",
                "Type",
                "Metric",
                "Expected",
                "Actual",
                "Deviation (σ)",
                "Severity",
                "Test ID",
            ]

            for col_idx, header in enumerate(anomaly_headers, 1):
                cell = anomalies_sheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, anomaly in enumerate(anomalies_response.anomalies, 2):
                anomalies_sheet.cell(
                    row=row_idx, column=1, value=anomaly.timestamp.isoformat()
                )
                anomalies_sheet.cell(
                    row=row_idx, column=2, value=anomaly.anomaly_type.value
                )
                anomalies_sheet.cell(row=row_idx, column=3, value=anomaly.metric_name)
                anomalies_sheet.cell(
                    row=row_idx, column=4, value=anomaly.expected_value
                )
                anomalies_sheet.cell(row=row_idx, column=5, value=anomaly.actual_value)
                anomalies_sheet.cell(
                    row=row_idx, column=6, value=anomaly.deviation_sigma
                )
                anomalies_sheet.cell(row=row_idx, column=7, value=anomaly.severity)
                anomalies_sheet.cell(row=row_idx, column=8, value=anomaly.test_id)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()


# ==================== Scheduled Reports Repository ====================


class ScheduledReportsRepository:
    """Repository for scheduled report configurations.

    Manages CRUD operations for scheduled report configurations
    stored in the analytics database.
    """

    def __init__(self, session: AsyncSession):
        """Initialize the repository.

        Args:
            session: SQLAlchemy async session.
        """
        self._session = session

    async def create_report(
        self, config: ScheduledReportConfig
    ) -> ScheduledReportConfig:
        """Create a new scheduled report configuration.

        Args:
            config: Report configuration.

        Returns:
            Created configuration with ID.
        """
        from atp.analytics.models import ScheduledReport

        report = ScheduledReport(
            name=config.name,
            frequency=config.frequency.value,
            recipients_json=config.recipients,
            include_trends=config.include_trends,
            include_anomalies=config.include_anomalies,
            include_correlations=config.include_correlations,
            filters_json=config.filters,
            is_active=config.is_active,
        )

        # Calculate next run time
        report.next_run = self._calculate_next_run(config.frequency)

        self._session.add(report)
        await self._session.flush()

        config.id = report.id
        config.next_run = report.next_run
        config.created_at = report.created_at

        return config

    async def get_report(self, report_id: int) -> ScheduledReportConfig | None:
        """Get a scheduled report by ID.

        Args:
            report_id: Report ID.

        Returns:
            Report configuration or None if not found.
        """
        from atp.analytics.models import ScheduledReport

        stmt = select(ScheduledReport).where(ScheduledReport.id == report_id)
        result = await self._session.execute(stmt)
        report = result.scalar_one_or_none()

        if report is None:
            return None

        return ScheduledReportConfig(
            id=report.id,
            name=report.name,
            frequency=ScheduledReportFrequency(report.frequency),
            recipients=report.recipients_json or [],
            include_trends=report.include_trends,
            include_anomalies=report.include_anomalies,
            include_correlations=report.include_correlations,
            filters=report.filters_json or {},
            is_active=report.is_active,
            last_run=report.last_run,
            next_run=report.next_run,
            created_at=report.created_at,
        )

    async def list_reports(
        self, *, is_active: bool | None = None
    ) -> list[ScheduledReportConfig]:
        """List scheduled reports.

        Args:
            is_active: Optional filter by active status.

        Returns:
            List of report configurations.
        """
        from atp.analytics.models import ScheduledReport

        stmt = select(ScheduledReport).order_by(ScheduledReport.name)

        if is_active is not None:
            stmt = stmt.where(ScheduledReport.is_active == is_active)

        result = await self._session.execute(stmt)
        reports = result.scalars().all()

        return [
            ScheduledReportConfig(
                id=r.id,
                name=r.name,
                frequency=ScheduledReportFrequency(r.frequency),
                recipients=r.recipients_json or [],
                include_trends=r.include_trends,
                include_anomalies=r.include_anomalies,
                include_correlations=r.include_correlations,
                filters=r.filters_json or {},
                is_active=r.is_active,
                last_run=r.last_run,
                next_run=r.next_run,
                created_at=r.created_at,
            )
            for r in reports
        ]

    async def update_report(
        self, report_id: int, updates: dict[str, Any]
    ) -> ScheduledReportConfig | None:
        """Update a scheduled report.

        Args:
            report_id: Report ID.
            updates: Fields to update.

        Returns:
            Updated configuration or None if not found.
        """
        from atp.analytics.models import ScheduledReport

        stmt = select(ScheduledReport).where(ScheduledReport.id == report_id)
        result = await self._session.execute(stmt)
        report = result.scalar_one_or_none()

        if report is None:
            return None

        for key, value in updates.items():
            if hasattr(report, key):
                if key == "frequency":
                    report.frequency = value.value if hasattr(value, "value") else value
                    report.next_run = self._calculate_next_run(
                        ScheduledReportFrequency(report.frequency)
                    )
                elif key == "recipients":
                    report.recipients_json = value
                elif key == "filters":
                    report.filters_json = value
                else:
                    setattr(report, key, value)

        await self._session.flush()

        return await self.get_report(report_id)

    async def delete_report(self, report_id: int) -> bool:
        """Delete a scheduled report.

        Args:
            report_id: Report ID.

        Returns:
            True if deleted, False if not found.
        """
        from atp.analytics.models import ScheduledReport

        stmt = select(ScheduledReport).where(ScheduledReport.id == report_id)
        result = await self._session.execute(stmt)
        report = result.scalar_one_or_none()

        if report is None:
            return False

        await self._session.delete(report)
        await self._session.flush()
        return True

    async def get_due_reports(self) -> list[ScheduledReportConfig]:
        """Get reports that are due to run.

        Returns:
            List of reports where next_run is in the past.
        """
        from atp.analytics.models import ScheduledReport

        now = datetime.now()
        stmt = (
            select(ScheduledReport)
            .where(ScheduledReport.is_active.is_(True))
            .where(ScheduledReport.next_run <= now)
            .order_by(ScheduledReport.next_run)
        )

        result = await self._session.execute(stmt)
        reports = result.scalars().all()

        return [
            ScheduledReportConfig(
                id=r.id,
                name=r.name,
                frequency=ScheduledReportFrequency(r.frequency),
                recipients=r.recipients_json or [],
                include_trends=r.include_trends,
                include_anomalies=r.include_anomalies,
                include_correlations=r.include_correlations,
                filters=r.filters_json or {},
                is_active=r.is_active,
                last_run=r.last_run,
                next_run=r.next_run,
                created_at=r.created_at,
            )
            for r in reports
        ]

    async def mark_report_run(self, report_id: int) -> None:
        """Mark a report as having been run.

        Updates last_run and calculates next_run.

        Args:
            report_id: Report ID.
        """
        from atp.analytics.models import ScheduledReport

        stmt = select(ScheduledReport).where(ScheduledReport.id == report_id)
        result = await self._session.execute(stmt)
        report = result.scalar_one_or_none()

        if report:
            report.last_run = datetime.now()
            report.next_run = self._calculate_next_run(
                ScheduledReportFrequency(report.frequency)
            )
            await self._session.flush()

    def _calculate_next_run(self, frequency: ScheduledReportFrequency) -> datetime:
        """Calculate the next run time based on frequency.

        Args:
            frequency: Report frequency.

        Returns:
            Next run datetime.
        """
        now = datetime.now()

        if frequency == ScheduledReportFrequency.DAILY:
            # Next day at 8:00 AM
            next_run = now.replace(hour=8, minute=0, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        elif frequency == ScheduledReportFrequency.WEEKLY:
            # Next Monday at 8:00 AM
            days_until_monday = (7 - now.weekday()) % 7
            if days_until_monday == 0 and now.hour >= 8:
                days_until_monday = 7
            next_run = (now + timedelta(days=days_until_monday)).replace(
                hour=8, minute=0, second=0, microsecond=0
            )
        else:  # monthly
            # First day of next month at 8:00 AM
            if now.month == 12:
                next_run = now.replace(
                    year=now.year + 1, month=1, day=1, hour=8, minute=0, second=0
                )
            else:
                next_run = now.replace(
                    month=now.month + 1, day=1, hour=8, minute=0, second=0
                )

        return next_run
